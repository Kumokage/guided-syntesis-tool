import marimo

__generated_with = "0.16.5"
app = marimo.App()


@app.cell(hide_code=True)
def _(mo):
    mo.md(r"""# Data Processing""")
    return


@app.cell
def _():
    excel_path = r"./data/titanium sols.xlsx"
    artifacts_path = r"./artifacts/second_dataset/"
    return artifacts_path, excel_path


@app.cell
def _(FILibExcel, excel_path):
    tables_dict = FILibExcel.get_all_tables(file_name=excel_path)
    df = tables_dict["Table2"]['dataframe']
    df = df.fillna(value=0)
    return (df,)


@app.cell
def _(df):
    df.shape
    return


@app.cell
def _(df):
    df.head()
    return


@app.cell
def _(df):
    df.drop('Composition mixtures', axis=1).astype('float').describe().T
    return


@app.cell
def _(df):
    composition_mixtures_dict = {'H2O + HNO3 + TiIPO + ButOH': 1, 'H2O + HNO3 + TiIPO + IPOОН': 2, 'H2O + HNO3 + ТiBut + ButOH': 3, 'H2O + TiOSO₄´xH2SO4´yH2O': 4}
    df_1 = df.replace(composition_mixtures_dict)
    return (df_1,)


@app.cell
def _(df_1):
    df_1.dtypes
    return


@app.cell
def _(df_1):
    X = df_1.drop(['Composition mixtures', 'Contents, %', 'd, nm', 'Stability of sols, days', 'ultrasound '], axis=1)
    y_days = df_1['Stability of sols, days']
    y_d = df_1['d, nm']
    y_content = df_1['Contents, %']
    return X, y_content, y_d, y_days


@app.cell
def _(X):
    X
    return


@app.cell(hide_code=True)
def _(mo):
    mo.md(r"""# Linear regression for coafitionts predictions""")
    return


@app.cell
def _(X, y_days):
    from sklearn.linear_model import LinearRegression
    from sklearn.model_selection import GridSearchCV
    _parameters = {'fit_intercept': [True, False]}
    _model = LinearRegression()
    _clf = GridSearchCV(_model, _parameters, cv=5, scoring='neg_mean_absolute_error', refit=True, n_jobs=-1)
    _clf.fit(X, y_days)
    best_linear_regression_days = _clf.best_estimator_
    print(_clf.best_score_)
    print(_clf.best_params_)
    print(best_linear_regression_days.coef_)
    return GridSearchCV, LinearRegression, best_linear_regression_days


@app.cell
def _(GridSearchCV, LinearRegression, X, y_d):
    from sklearn.preprocessing import StandardScaler
    from sklearn.pipeline import make_pipeline
    _parameters = {'fit_intercept': [True, False]}
    _model = LinearRegression()
    _clf = GridSearchCV(_model, _parameters, cv=5, scoring='neg_mean_absolute_error', refit=True, n_jobs=-1)
    pipe = make_pipeline(StandardScaler(), _clf)
    pipe.fit(X, y_d)
    best_linear_regression_d = _clf.best_estimator_
    print(_clf.best_score_)
    print(_clf.best_params_)
    print(best_linear_regression_d.coef_)
    return (best_linear_regression_d,)


@app.cell
def _(GridSearchCV, LinearRegression, X, y_content):
    _parameters = {'fit_intercept': [True, False]}
    _model = LinearRegression()
    _clf = GridSearchCV(_model, _parameters, cv=5, scoring='neg_mean_absolute_error', refit=True, n_jobs=-1)
    _clf.fit(X, y_content)
    best_linear_regression_content = _clf.best_estimator_
    print(_clf.best_score_)
    print(_clf.best_params_)
    print(best_linear_regression_content.coef_)
    return (best_linear_regression_content,)


@app.cell(hide_code=True)
def _(mo):
    mo.md(r"""# Optimisation with NSGA2 algorithm""")
    return


@app.cell
def _():
    # (use marimo's built-in package management features instead) !pip install -U pymoo
    return


@app.cell
def _(best_linear_regression_days, np):
    -np.append(best_linear_regression_days.coef_, best_linear_regression_days.intercept_)
    return


@app.cell
def _(
    best_linear_regression_content,
    best_linear_regression_d,
    best_linear_regression_days,
    np,
):
    from pymoo.algorithms.moo.nsga2 import NSGA2
    from pymoo.core.problem import ElementwiseProblem
    from pymoo.optimize import minimize
    from pymoo.visualization.scatter import Scatter

    class SoleProblem(ElementwiseProblem):

        def __init__(self, coefs):
            self.coefs = coefs
            super().__init__(n_var=4, n_obj=3, n_ieq_constr=4, xl=np.array([20, 1, 0.1, 0.14]), xu=np.array([80, 130, 5, 1.5]))

        def _evaluate(self, x, out, *args, **kwargs):
            fs = []
            gs = []
            for coef in self.coefs:
                f = coef[-1]
                for i in range(0, len(coef) - 1):
                    f = f + x[i] * coef[i]
                fs.append(f)
            out['F'] = fs
            out['G'] = [fs[0], -fs[1] + 0.7, fs[2], -fs[2] - 100]
    problem = SoleProblem([-np.append(best_linear_regression_days.coef_, best_linear_regression_days.intercept_), np.append(best_linear_regression_d.coef_, best_linear_regression_d.intercept_), -np.append(best_linear_regression_content.coef_, best_linear_regression_content.intercept_)])
    algorithm = NSGA2(pop_size=100)
    res = minimize(problem, algorithm, ('n_gen', 500), verbose=False, seed=1)
    return (res,)


@app.cell
def _(artifacts_path, res):
    import plotly.express as px
    F = res.F
    _fig = px.scatter(x=-F[:, 0], y=F[:, 1], labels={'x': 'Stability of sols, days', 'y': 'd, nm'})
    _fig.write_html(artifacts_path + 'first_figure.html', auto_open=True)
    _fig.write_image(artifacts_path + 'optimal.png')
    return


@app.cell
def _(res):
    for f, x in zip(res.F, res.X):
        print(f'For getting stability days = {round(-f[0], 3)}, diameter = {round(f[1], 3)} and content = {round(-f[2], 3)} use:')
        print(f'\ttemperature = {round(x[0], 3)}')
        print(f'\ttime = {round(x[1], 3)}')
        print(f'\tc(acid) = {round(x[2], 3)}')
        print(f'\tc(Ti) = {round(x[3], 3)}')
    return


@app.cell(hide_code=True)
def _(mo):
    mo.md(r"""# Random forest feature importance prediction""")
    return


@app.cell
def _():
    # (use marimo's built-in package management features instead) !pip install xgboost
    return


@app.cell
def _(df_1):
    df_2 = df_1.rename(columns={'t, °С': 'T, °С'})
    df_2 = df_2.rename(columns={'с(Ti4+), M': 'с(Ti$^{4+}$), M'})
    df_2 = df_2.rename(columns={'Composition mixtures': 'Mixture composition'})
    return (df_2,)


@app.cell
def _(df_2):
    X_1 = df_2.drop(['Contents, %', 'd, nm', 'Stability of sols, days', 'ultrasound '], axis=1)
    y_days_1 = df_2['Stability of sols, days']
    y_d_1 = df_2['d, nm']
    y_content_1 = df_2['Contents, %']
    return X_1, y_content_1, y_d_1, y_days_1


@app.cell
def _(X_1):
    X_1
    return


@app.cell
def _():
    from sklearn.ensemble import RandomForestRegressor, GradientBoostingRegressor
    import xgboost as xgb
    return GradientBoostingRegressor, RandomForestRegressor, xgb


@app.cell
def _(np):
    def hex_to_RGB(hex_str):
        """ #FFFFFF -> [255,255,255]"""
        #Pass 16 to the integer function for change of base
        return [int(hex_str[i:i+2], 16) for i in range(1,6,2)]

    def get_color_gradient(c1, c2, n):
        """
        Given two hex colors, returns a color gradient
        with n colors.
        """
        assert n > 1
        c1_rgb = np.array(hex_to_RGB(c1))/255
        c2_rgb = np.array(hex_to_RGB(c2))/255
        mix_pcts = [x/(n-1) for x in range(n)]
        rgb_colors = [((1-mix)*c1_rgb + (mix*c2_rgb)) for mix in mix_pcts]
        return ["#" + "".join([format(int(round(val*255)), "02x") for val in item]) for item in rgb_colors]
    return (get_color_gradient,)


@app.cell
def _(X_1, artifacts_path, get_color_gradient, np, plt):
    from matplotlib.offsetbox import AnchoredText

    def draw_importance(importances, model_name, columns=X_1.columns, ax=None, is_save=False):
        features = {}
        color1 = '#2D466D'
        color2 = '#A2B0C5'
        for i, feature in enumerate(columns):
            features[f'f{i + 1}'] = feature
        indices = np.argsort(importances)[::-1]
        num_to_plot = len(columns)
        feature_indices = [ind + 1 for ind in indices[:num_to_plot]]
        print('Feature ranking:')
        for f in range(num_to_plot):
            print('%d. %s %f ' % (f + 1, features['f' + str(feature_indices[f])], importances[indices[f]]))
        is_save = False
        if ax is None:
            plt.figure(figsize=(20, 10))
            ax = plt.gca()
            is_save = True
        bars = ax.bar(range(num_to_plot), importances[indices[:num_to_plot]], color=get_color_gradient(color1, color2, num_to_plot), align='center')
        ax.set_xticks(range(num_to_plot), feature_indices)
        ax.set_xlim([-1, num_to_plot])
        ax.legend(bars, [u''.join(features['f' + str(i)]) for i in feature_indices], fontsize=20)
        ax.tick_params(axis='both', labelsize=22)
        if is_save:
            plt.savefig(f'{artifacts_path}{model_name}.png', format='png', dpi=600, bbox_inches='tight', pad_inches=0.1)
            plt.show()
        else:
            txt = AnchoredText(model_name, loc='lower right', pad=0.25, borderpad=0, prop=dict(fontsize='xx-large'))
            ax.add_artist(txt)  # ax.set_title(f"Feature importance in {model_name}", fontsize=18)
    return AnchoredText, draw_importance


@app.cell
def _(GridSearchCV, RandomForestRegressor, draw_importance):
    def draw_random_forest(X, y, name, ax=None, is_save=False):
        _parameters = {'n_estimators': range(100, 500, 100), 'max_depth': [None] + list(range(3, 11, 2))}
        _model = RandomForestRegressor(random_state=42)
        _clf = GridSearchCV(_model, _parameters, cv=5, scoring='neg_mean_absolute_error', refit=True, n_jobs=-1)
        _clf.fit(X, y)
        best_random_forest = _clf.best_estimator_
        print(_clf.best_score_)
        print(_clf.best_params_)
        draw_importance(best_random_forest.feature_importances_, f'{name}', ax=ax, is_save=is_save)
    return (draw_random_forest,)


@app.cell
def _(GradientBoostingRegressor, GridSearchCV, draw_importance):
    def draw_gradient_boosting(X, y, name, ax=None, is_save=False):
        _parameters = {'learning_rate': [0.5, 0.25, 0.1, 0.05, 0.01], 'n_estimators': [4, 8, 16, 32, 64, 128], 'max_depth': range(1, 18, 2)}
        _model = GradientBoostingRegressor(random_state=42)
        _clf = GridSearchCV(_model, _parameters, cv=5, scoring='neg_mean_absolute_error', refit=True, n_jobs=-1)
        _clf.fit(X, y)
        best_gradient_boost = _clf.best_estimator_
        print(_clf.best_score_)
        print(_clf.best_params_)
        draw_importance(best_gradient_boost.feature_importances_, f'{name}', ax=ax, is_save=is_save)
    return (draw_gradient_boosting,)


@app.cell
def _(GridSearchCV, draw_importance, xgb):
    def draw_xgboost(X, y, name, ax=None, is_save=False):
        _parameters = {'min_child_weight': [1, 5, 7, 10], 'gamma': [0.5, 1, 1.5, 2, 2.5], 'subsample': [0.6, 0.8, 1.0], 'colsample_bytree': [0.6, 0.8, 1.0], 'max_depth': [3, 4, 5]}
        _model = xgb.XGBRegressor(learning_rate=0.02, n_estimators=600, nthread=1, seed=0)
        _clf = GridSearchCV(_model, _parameters, cv=5, scoring='neg_mean_absolute_error', refit=True, n_jobs=-1)
        _clf.fit(X, y)
        best_xgboost = _clf.best_estimator_
        print(_clf.best_score_)
        print(_clf.best_params_)
        draw_importance(best_xgboost.feature_importances_, f'{name}', ax=ax, is_save=is_save)
    return (draw_xgboost,)


@app.cell(hide_code=True)
def _(mo):
    mo.md(r"""## Stability days""")
    return


@app.cell
def _():
    is_save = False
    return (is_save,)


@app.cell
def _(
    X_1,
    artifacts_path,
    draw_gradient_boosting,
    draw_random_forest,
    draw_xgboost,
    is_save,
    plt,
    y_days_1,
):
    if is_save:
        draw_random_forest(X_1, y_days_1, 'rf days of stability', is_save=is_save)
        draw_gradient_boosting(X_1, y_days_1, 'gb days of stability', is_save=is_save)
        draw_xgboost(X_1, y_days_1, 'xgb days of stability', is_save=is_save)
    else:
        _fig, _axes = plt.subplots(ncols=3, nrows=1, figsize=(24, 7))
        _fig.tight_layout(pad=3.5)
        draw_random_forest(X_1, y_days_1, 'a', ax=_axes[0])
        draw_gradient_boosting(X_1, y_days_1, 'b', ax=_axes[1])
        draw_xgboost(X_1, y_days_1, 'c', ax=_axes[2])
        plt.savefig(f'{artifacts_path}_StabilityDays.png', format='png', dpi=600, bbox_inches='tight', pad_inches=0)
    return


@app.cell(hide_code=True)
def _(mo):
    mo.md(r"""## Diameter""")
    return


@app.cell
def _(
    X_1,
    artifacts_path,
    draw_gradient_boosting,
    draw_random_forest,
    draw_xgboost,
    is_save,
    plt,
    y_d_1,
):
    if is_save:
        draw_random_forest(X_1, y_d_1, 'rf diameter', is_save=is_save)
        draw_gradient_boosting(X_1, y_d_1, 'gb diameter', is_save=is_save)
        draw_xgboost(X_1, y_d_1, 'xgb diameter', is_save=is_save)
    else:
        _fig, _axes = plt.subplots(ncols=3, nrows=1, figsize=(24, 7))
        _fig.tight_layout(pad=3.5)
        draw_random_forest(X_1, y_d_1, 'a', ax=_axes[0])
        draw_gradient_boosting(X_1, y_d_1, 'b', ax=_axes[1])
        draw_xgboost(X_1, y_d_1, 'c', ax=_axes[2])
        plt.savefig(f'{artifacts_path}_Diameter.png', format='png', dpi=600, bbox_inches='tight', pad_inches=0.1)
    return


@app.cell(hide_code=True)
def _(mo):
    mo.md(r"""## Content""")
    return


@app.cell
def _(
    X_1,
    artifacts_path,
    draw_gradient_boosting,
    draw_random_forest,
    draw_xgboost,
    is_save,
    plt,
    y_content_1,
):
    if is_save:
        draw_random_forest(X_1, y_content_1, 'rf content', is_save=is_save)
        draw_gradient_boosting(X_1, y_content_1, 'gb content', is_save=is_save)
        draw_xgboost(X_1, y_content_1, 'xgb content', is_save=is_save)
    else:
        _fig, _axes = plt.subplots(ncols=3, nrows=1, figsize=(24, 7))
        _fig.tight_layout(pad=3.5)
        draw_random_forest(X_1, y_content_1, 'a', ax=_axes[0])
        draw_gradient_boosting(X_1, y_content_1, 'b', ax=_axes[1])
        draw_xgboost(X_1, y_content_1, 'c', ax=_axes[2])
        plt.savefig(f'{artifacts_path}_Content.png', format='png', dpi=600, bbox_inches='tight', pad_inches=0)
    return


@app.cell(hide_code=True)
def _(mo):
    mo.md(r"""# Results""")
    return


@app.cell
def _(df_2):
    new_col = []
    for elem in df_2['d, nm']:
        if elem > 2000:
            new_col.append(5)
        elif elem > 1000:
            new_col.append(4)
        elif elem > 400:
            new_col.append(3)
        elif elem > 2:
            new_col.append(2)
        elif elem > 1:
            new_col.append(1)
        else:
            new_col.append(0)
    df_2['d_categorical'] = new_col
    return


@app.cell
def _(artifacts_path, df_2, plt, sns):
    sns.set(font_scale=1.4)
    _fig, ax = plt.subplots()
    _fig.set_size_inches(10, 7)
    ax.tick_params(axis='both', labelsize=16)
    sns.histplot(df_2, x='с(Ti$^{4+}$), M', hue='d_categorical', multiple='stack', palette='mako')
    plt.show()
    plt.savefig(artifacts_path + 'Ti4Distribution.png', format='png', dpi=600)
    plt.show()
    return


@app.cell
def _(AnchoredText, artifacts_path, plt, sns):
    def draw_hists(df, hue, name, hue_order=None, discrete_save=False):
        fontdict = {'weight': 'bold', 'size': 20}
        _fig, _axes = plt.subplots(ncols=2, nrows=3)
        _fig.set_size_inches(20, 25)
        sns.set(font_scale=1.4)
        titles = ['a', 'b', 'c', 'd', 'e']
        index = 0
        if discrete_save:
            plt.show()
        for col, ax in zip(['T, °С', 't, min', 'с(acid), M', 'с(Ti$^{4+}$), M', 'Mixture composition'], _axes.flat):
            if discrete_save:
                plt.gcf().set_size_inches(10, 7)
                ax = plt.subplot()
            else:
                txt = AnchoredText(titles[index], loc='lower right', pad=0.25, borderpad=0, prop=dict(fontsize='large'))
                ax.add_artist(txt)
            ax.tick_params(axis='both', labelsize=20, width=5)
            ax.set_xlabel(col, fontdict=fontdict)
            ax.set_ylabel('Count', fontdict=fontdict)
            if col == 'Mixture composition':
                ax.set_xticks([1, 2, 3, 4])
                sns.histplot(df, x=col, hue=hue, hue_order=_hue_order, multiple='stack', palette='mako', ax=ax, discrete=True, shrink=0.7)
            else:
                sns.histplot(df, x=col, hue=hue, hue_order=_hue_order, multiple='stack', palette='mako', ax=ax)
            index = index + 1
            if discrete_save:
                plt.savefig(artifacts_path + f'second_stage_{name}_{col}.png', format='png', dpi=600)
                plt.show()
        _fig.subplots_adjust(hspace=0.15)
        _axes[-1, -1].axis('off')
        plt.savefig(artifacts_path + f'second_stage_{name}.png', format='png', dpi=600)
        plt.show()
    return (draw_hists,)


@app.cell
def _(df_2, draw_hists):
    draw_hists(df_2, 'd_categorical', 'distibution_with_d_categorical', discrete_save=True)
    draw_hists(df_2, 'd_categorical', 'distibution_with_d_categorical')
    return


@app.cell
def _(df_2):
    df_plot = df_2.copy(deep=True)

    def _replace_func_days(x):
        if x < 20:
            return '< 20'
        else:
            return str(x)
    df_plot['Stability of sols, days'] = df_plot['Stability of sols, days'].apply(_replace_func_days)
    return (df_plot,)


@app.cell
def _(df_plot, draw_hists):
    _hue_order = ['90.0', '60.0', '40.0', '< 20']
    draw_hists(df_plot, 'Stability of sols, days', 'distibution_with_stability_days', hue_order=_hue_order, discrete_save=True)
    draw_hists(df_plot, 'Stability of sols, days', 'distibution_with_stability_days', hue_order=_hue_order)
    return


@app.cell
def _(df_2):
    df_plot_1 = df_2.copy(deep=True)

    def _replace_func_days(x):
        if x < 96:
            return '< 96'
        else:
            return str(x)
    df_plot_1['Contents, %'] = df_plot_1['Contents, %'].apply(_replace_func_days)
    return (df_plot_1,)


@app.cell
def _(df_plot_1, draw_hists):
    _hue_order = ['100.0', '99.7', '99.5', '99.3', '98.8', '97.8', '97.5', '96.6', '< 96']
    draw_hists(df_plot_1, 'Contents, %', 'distibution_with_contents', hue_order=_hue_order, discrete_save=True)
    draw_hists(df_plot_1, 'Contents, %', 'distibution_with_contents', hue_order=_hue_order)
    return


@app.cell
def _():
    import marimo as mo
    import numpy as np
    import pandas as pd
    import openpyxl as xl
    import FILibExcel
    import seaborn as sns
    import matplotlib.pyplot as plt
    return FILibExcel, mo, np, plt, sns


if __name__ == "__main__":
    app.run()
