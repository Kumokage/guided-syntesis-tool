import marimo

__generated_with = "0.15.3"
app = marimo.App(width="medium")


@app.cell(hide_code=True)
def _(mo):
    mo.md(r"""# Imports""")
    return


@app.cell
def _():
    # (use marimo's built-in package management features instead) !pip install plotly==5.14.1
    # (use marimo's built-in package management features instead) !pip install "jupyterlab>=3" "ipywidgets>=7.6"
    # (use marimo's built-in package management features instead) !pip install jupyter-dash
    # (use marimo's built-in package management features instead) !pip install -U kaleido
    return


@app.cell
def _():
    # (use marimo's built-in package management features instead) !pip install numpy
    # (use marimo's built-in package management features instead) !pip install pandas
    # (use marimo's built-in package management features instead) !pip install openpyxl
    # (use marimo's built-in package management features instead) !pip install seaborn
    # (use marimo's built-in package management features instead) !pip install matplotlib
    # (use marimo's built-in package management features instead) !pip install -U scikit-learn
    # (use marimo's built-in package management features instead) !pip install shap
    # (use marimo's built-in package management features instead) !pip install xgboost
    return


@app.cell
def _():
    import numpy as np
    import pandas as pd
    import openpyxl as xl
    import FILibExcel
    import seaborn as sns
    import matplotlib.pyplot as plt
    from sklearn.linear_model import LinearRegression
    from sklearn.model_selection import GridSearchCV
    from sklearn.preprocessing import StandardScaler
    from sklearn.pipeline import make_pipeline
    from sklearn.ensemble import RandomForestRegressor, GradientBoostingRegressor
    import xgboost as xgb
    import shap

    return (
        FILibExcel,
        GradientBoostingRegressor,
        GridSearchCV,
        RandomForestRegressor,
        np,
        plt,
        shap,
        xgb,
    )


@app.cell(hide_code=True)
def _(mo):
    mo.md(r"""# Data Processing""")
    return


@app.cell
def _():
    excel_path = r"../../static/mock/titanium sols.xlsx"
    artifacts_path = r"../../static/mock/soles_artifacts/third_dataset/"
    return artifacts_path, excel_path


@app.cell
def _(FILibExcel, excel_path):
    tables_dict = FILibExcel.get_all_tables(file_name=excel_path)
    df = tables_dict["Table3"]["dataframe"]
    df = df.fillna(value=0)
    return (df,)


@app.cell
def _(df):
    df.shape
    return


@app.cell
def _(df):
    df.head()
    return


@app.cell
def _(df):
    df.drop("Composition mixtures", axis=1).astype("float").describe().T
    return


@app.cell
def _(df):
    df.dtypes
    return


@app.cell
def _(df):
    composition_mixtures_dict = {
        "H2O + HNO3 + TiIPO + ButOH": 1,
        "H2O + HNO3 + TiIPO + IPOОН": 2,
        "H2O + HNO3 + ТiBut + ButOH": 3,
        "H2O + TiOSO₄´xH2SO4´yH2O": 4,
    }
    df_1 = df.replace(composition_mixtures_dict)
    return (df_1,)


@app.cell
def _(df_1):
    df_2 = df_1.rename(columns={"t, °С": "T, °С"})
    df_2 = df_2.rename(columns={"с(Ti4+), M": "с(Ti$^{4+}$), M"})
    df_2 = df_2.rename(columns={"Composition mixtures": "Mixture composition"})
    X = df_2.drop(["Contents, %", "d, nm", "Stability of sols, days"], axis=1)
    y_days = df_2["Stability of sols, days"]
    y_d = df_2["d, nm"]
    y_content = df_2["Contents, %"]
    return X, y_content, y_d, y_days


@app.cell
def _(X):
    X
    return


@app.cell(hide_code=True)
def _(mo):
    mo.md(r"""# Random forest feature importance prediction""")
    return


@app.cell(hide_code=True)
def _(mo):
    mo.md(r"""## Useful functions""")
    return


@app.cell
def _(np):
    def hex_to_RGB(hex_str):
        """#FFFFFF -> [255,255,255]"""
        return [int(hex_str[i : i + 2], 16) for i in range(1, 6, 2)]

    def get_color_gradient(c1, c2, n):
        """
        Given two hex colors, returns a color gradient
        with n colors.
        """
        assert n > 1
        c1_rgb = np.array(hex_to_RGB(c1)) / 255
        c2_rgb = np.array(hex_to_RGB(c2)) / 255
        mix_pcts = [_x / (n - 1) for _x in range(n)]
        rgb_colors = [(1 - mix) * c1_rgb + mix * c2_rgb for mix in mix_pcts]
        return [
            "#" + "".join([format(int(round(val * 255)), "02x") for val in item])
            for item in rgb_colors
        ]

    return (get_color_gradient,)


@app.cell
def _(X, artifacts_path, get_color_gradient, np, plt):
    def draw_importance(importances, model_name, columns=X.columns):
        features = {}
        color1 = "#2D466D"
        color2 = "#A2B0C5"
        for i, feature in enumerate(columns):
            features[f"f{i+1}"] = feature

        indices = np.argsort(importances)[::-1]
        num_to_plot = len(columns)
        feature_indices = [ind + 1 for ind in indices[:num_to_plot]]

        print("Feature ranking:")
        for f in range(num_to_plot):
            print(
                "%d. %s %f "
                % (
                    f + 1,
                    features["f" + str(feature_indices[f])],
                    importances[indices[f]],
                )
            )
        plt.figure(figsize=(20, 10))
        bars = plt.bar(
            range(num_to_plot),
            importances[indices[:num_to_plot]],
            color=get_color_gradient(color1, color2, num_to_plot),
            align="center",
        )
        ticks = plt.xticks(range(num_to_plot), feature_indices, fontsize=22)
        plt.yticks(fontsize=22)
        plt.xlim([-1, num_to_plot])
        plt.legend(
            bars,
            ["".join(features["f" + str(i)]) for i in feature_indices],
            fontsize="24",
        )
        plt.title(f"Feature importance in {model_name}", fontsize=22)
        plt.savefig(
            f"{artifacts_path}{model_name}.eps",
            format="eps",
            dpi=600,
            bbox_inches="tight",
            transparent="True",
            pad_inches=0,
        )

    return (draw_importance,)


@app.cell
def _(GridSearchCV, RandomForestRegressor, draw_importance):
    def learn_random_forest(X, y, name):
        parameters = {
            "n_estimators": range(100, 500, 100),
            "max_depth": [None] + list(range(3, 11, 2)),
        }
        model = RandomForestRegressor(random_state=42)
        clf = GridSearchCV(
            model,
            parameters,
            cv=5,
            scoring="neg_mean_absolute_error",
            refit=True,
            n_jobs=-1,
        )
        clf.fit(X, y)
        best_random_forest = clf.best_estimator_
        print(clf.best_score_)
        print(clf.best_params_)
        draw_importance(
            best_random_forest.feature_importances_, f"Random forest {name}"
        )
        return best_random_forest

    return (learn_random_forest,)


@app.cell
def _(GradientBoostingRegressor, GridSearchCV, draw_importance):
    def learn_gradient_boosting(X, y, name):
        parameters = {
            "learning_rate": [0.5, 0.25, 0.1, 0.05, 0.01],
            "n_estimators": [4, 8, 16, 32, 64, 128],
            "max_depth": range(1, 18, 2),
        }
        model = GradientBoostingRegressor(random_state=42)
        clf = GridSearchCV(
            model,
            parameters,
            cv=5,
            scoring="neg_mean_absolute_error",
            refit=True,
            n_jobs=-1,
        )
        clf.fit(X, y)
        best_gradient_boost = clf.best_estimator_
        print(clf.best_score_)
        print(clf.best_params_)
        draw_importance(
            best_gradient_boost.feature_importances_, f"Gradient boost {name}"
        )
        return best_gradient_boost

    return (learn_gradient_boosting,)


@app.cell
def _(GridSearchCV, draw_importance, xgb):
    def learn_xgboost(X, y, name):
        parameters = {
            "min_child_weight": [1, 5, 7, 10],
            "gamma": [0.5, 1, 1.5, 2, 2.5],
            "subsample": [0.6, 0.8, 1.0],
            "colsample_bytree": [0.6, 0.8, 1.0],
            "max_depth": [3, 4, 5],
        }
        model = xgb.XGBRegressor(
            learning_rate=0.02, n_estimators=600, nthread=1, seed=0
        )
        clf = GridSearchCV(
            model,
            parameters,
            cv=5,
            scoring="neg_mean_absolute_error",
            refit=True,
            n_jobs=-1,
        )
        clf.fit(X, y)
        best_xgboost = clf.best_estimator_
        print(clf.best_score_)
        print(clf.best_params_)
        draw_importance(best_xgboost.feature_importances_, f"XGboost {name}")
        return best_xgboost

    return (learn_xgboost,)


@app.cell(hide_code=True)
def _(mo):
    mo.md(r"""## Learning""")
    return


@app.cell
def _(X, learn_random_forest, y_days):
    days_rf = learn_random_forest(X, y_days, "days of stability")
    return (days_rf,)


@app.cell
def _(X, learn_gradient_boosting, y_days):
    days_grad = learn_gradient_boosting(X, y_days, "days of stability")
    return (days_grad,)


@app.cell
def _(X, learn_xgboost, y_days):
    days_xgboost = learn_xgboost(X, y_days, "days of stability")
    return (days_xgboost,)


@app.cell
def _(X, learn_random_forest, y_d):
    d_rf = learn_random_forest(X, y_d, "diameter")
    return (d_rf,)


@app.cell
def _(X, learn_gradient_boosting, y_d):
    d_grad = learn_gradient_boosting(X, y_d, "diameter")
    return (d_grad,)


@app.cell
def _(X, learn_xgboost, y_d):
    d_xgboost = learn_xgboost(X, y_d, "diameter")
    return (d_xgboost,)


@app.cell
def _(X, learn_random_forest, y_content):
    content_rf = learn_random_forest(X, y_content, "content")
    return (content_rf,)


@app.cell
def _(X, learn_gradient_boosting, y_content):
    content_gb = learn_gradient_boosting(X, y_content, "content")
    return (content_gb,)


@app.cell
def _(X, learn_xgboost, y_content):
    content_xgb = learn_xgboost(X, y_content, "content")
    return (content_xgb,)


@app.cell(hide_code=True)
def _(mo):
    mo.md(r"""## Shap""")
    return


@app.cell
def _(X, artifacts_path, plt, shap):
    def explain(model, X=X, name="model", is_save=False, color_bar=False):
        X = X.rename(columns={"Mixture composition": "Com. mix."})
        explainer = shap.Explainer(model, X)
        shap_values = explainer(X)
        shap.plots.beeswarm(
            shap_values, plot_size=[10, 6], s=26, show=False, color_bar=color_bar
        )
        if is_save:
            plt.tick_params(axis="both", labelsize=16)
            plt.savefig(f"{artifacts_path}{name} shap.png", bbox_inches="tight")
            plt.show()

    return (explain,)


@app.cell(hide_code=True)
def _(mo):
    mo.md(r"""### Days of stability""")
    return


@app.cell
def _(days_grad, days_rf, days_xgboost, explain):
    explain(
        days_xgboost, name="XGBoost, days of stability", color_bar=True, is_save=True
    )
    explain(
        days_grad,
        name="GradientBoosting, days of stability",
        color_bar=True,
        is_save=True,
    )
    explain(
        days_rf, name="RandomForest, days of stability", color_bar=True, is_save=True
    )
    return


@app.cell
def _(artifacts_path, days_grad, days_rf, days_xgboost, explain, plt):
    from matplotlib.offsetbox import AnchoredText

    plt.figure(figsize=(20, 20))
    plt.subplot(3, 1, 1)
    _y = 0.02
    _x = 1.03
    _txt = AnchoredText(
        "a", loc="upper left", pad=0.25, borderpad=-0.5, prop=dict(fontsize="xx-large")
    )
    plt.gca().add_artist(_txt)
    explain(days_rf, name="RandomForest, days of stability", color_bar=True)
    plt.tick_params(axis="both", labelsize=16)
    plt.subplot(3, 1, 2)
    _txt = AnchoredText(
        "b", loc="upper left", pad=0.25, borderpad=-0.5, prop=dict(fontsize="xx-large")
    )
    plt.gca().add_artist(_txt)
    explain(days_grad, name="GradientBoosting, days of stability", color_bar=True)
    plt.tick_params(axis="both", labelsize=16)
    plt.subplot(3, 1, 3)
    _txt = AnchoredText(
        "c", loc="upper left", pad=0.25, borderpad=-0.5, prop=dict(fontsize="xx-large")
    )
    plt.gca().add_artist(_txt)
    explain(days_xgboost, name="XGBoost, days of stability", color_bar=True)
    plt.subplots_adjust(top=2.5)
    plt.tick_params(axis="both", labelsize=16)
    plt.savefig(f"{artifacts_path}Stability shap.png", bbox_inches="tight")
    return (AnchoredText,)


@app.cell(hide_code=True)
def _(mo):
    mo.md(r"""### Diameter of zole""")
    return


@app.cell
def _(d_grad, d_rf, d_xgboost, explain):
    explain(d_rf, name="RandomForest, diameter", color_bar=True, is_save=True)
    explain(d_grad, name="GradientBoosting, diameter", color_bar=True, is_save=True)
    explain(d_xgboost, name="XGBoost, diameter", color_bar=True, is_save=True)
    return


@app.cell
def _(AnchoredText, artifacts_path, d_grad, d_rf, d_xgboost, explain, plt):
    plt.figure(figsize=(20, 7))  # width large, height smaller to fit horizontally

    plt.subplot(1, 3, 1)
    _txt = AnchoredText(
        "a", loc="upper left", pad=0.25, borderpad=-0.5, prop=dict(fontsize="xx-large")
    )
    plt.gca().add_artist(_txt)
    explain(d_rf, name="RandomForest, diameter", color_bar=True)
    plt.tick_params(axis="both", labelsize=16)

    plt.subplot(1, 3, 2)
    _txt = AnchoredText(
        "b", loc="upper left", pad=0.25, borderpad=-0.5, prop=dict(fontsize="xx-large")
    )
    plt.gca().add_artist(_txt)
    explain(d_grad, name="GradientBoosting, diameter", color_bar=True)
    plt.tick_params(axis="both", labelsize=16)

    plt.subplot(1, 3, 3)
    _txt = AnchoredText(
        "c", loc="upper left", pad=0.25, borderpad=-0.5, prop=dict(fontsize="xx-large")
    )
    plt.gca().add_artist(_txt)
    explain(d_xgboost, name="XGBoost, diameter", color_bar=True)
    plt.tick_params(axis="both", labelsize=16)

    plt.subplots_adjust(wspace=0.2, right=2.5)  # add some horizontal spacing
    plt.savefig(f"{artifacts_path}Diameter shap.eps", bbox_inches="tight")
    return


@app.cell(hide_code=True)
def _(mo):
    mo.md(r"""### Content""")
    return


@app.cell
def _(content_gb, content_rf, content_xgb, explain):
    explain(content_rf, name="RandomForest, content", color_bar=True, is_save=True)
    explain(content_gb, name="GradientBoosting, content", color_bar=True, is_save=True)
    explain(content_xgb, name="XGBoost, content", color_bar=True, is_save=True)
    return


@app.cell
def _(
    AnchoredText,
    artifacts_path,
    content_gb,
    content_rf,
    content_xgb,
    explain,
    plt,
):
    plt.figure(figsize=(20, 20))
    plt.subplot(3, 1, 1)
    _y = 0.02
    _x = 1.03
    _txt = AnchoredText(
        "a", loc="upper left", pad=0.25, borderpad=-0.5, prop=dict(fontsize="xx-large")
    )
    plt.gca().add_artist(_txt)
    explain(content_rf, name="RandomForest, content", color_bar=True)
    plt.tick_params(axis="both", labelsize=16)
    plt.subplot(3, 1, 2)
    _txt = AnchoredText(
        "b", loc="upper left", pad=0.25, borderpad=-0.5, prop=dict(fontsize="xx-large")
    )
    plt.gca().add_artist(_txt)
    explain(content_gb, name="GradientBoosting, content", color_bar=True)
    plt.tick_params(axis="both", labelsize=16)
    plt.subplot(3, 1, 3)
    _txt = AnchoredText(
        "c", loc="upper left", pad=0.25, borderpad=-0.5, prop=dict(fontsize="xx-large")
    )
    plt.gca().add_artist(_txt)
    explain(content_xgb, name="XGBoost, content", color_bar=True)
    plt.subplots_adjust(top=2.5)
    plt.tick_params(axis="both", labelsize=16)
    plt.savefig(f"{artifacts_path}Content shap.png", bbox_inches="tight")
    return


@app.cell
def _():
    import marimo as mo

    return (mo,)


if __name__ == "__main__":
    app.run()
